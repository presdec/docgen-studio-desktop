import json
import re
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

PLACEHOLDER_RE = re.compile(r"\{\{([A-Z0-9_]+)\}\}")


def normalize_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        normalized = format(Decimal(str(value)).normalize(), "f")
        return normalized.rstrip("0").rstrip(".") if "." in normalized else normalized
    return str(value).strip()


def render_template(template_text, values):
    def replacer(match):
        token = match.group(1)
        return values.get(token, match.group(0))

    return PLACEHOLDER_RE.sub(replacer, template_text)


def build_row_values(worksheet, row_number, mapping):
    values = {}
    for token, column in mapping.items():
        values[token] = normalize_cell_value(worksheet[f"{column}{row_number}"].value)
    return values


def write_email_drafts(path, drafts):
    blocks = []
    for row_number, rendered_text in drafts:
        blocks.append(f"===== ROW {row_number} =====")
        blocks.append(rendered_text.rstrip())
        blocks.append("")
    path.write_text("\n".join(blocks).rstrip() + "\n", encoding="utf-8")


def write_report(path, payload, placeholders, generated_rows, skipped_rows):
    lines = [
        "Email Draft Generation Report",
        "",
        f"Workbook: {payload['workbook_path']}",
        f"Worksheet: {payload['worksheet_name']}",
        f"Output directory: {payload['output_dir']}",
        "",
        f"Email placeholders: {', '.join(sorted(placeholders)) or '(none)'}",
        f"Generated rows: {len(generated_rows)}",
        f"Skipped rows: {len(skipped_rows)}",
        "",
        "Skipped rows:",
    ]

    if skipped_rows:
        for row_number, missing in skipped_rows:
            lines.append(f"- Row {row_number}: missing {', '.join(missing)}")
    else:
        lines.append("- None")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    payload = json.loads(sys.argv[1])

    workbook = load_workbook(payload["workbook_path"], data_only=True)
    worksheet_name = payload["worksheet_name"]
    if worksheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Worksheet {worksheet_name!r} not found. Available sheets: {', '.join(workbook.sheetnames)}"
        )

    worksheet = workbook[worksheet_name]
    data_start_row = int(payload["data_start_row"])
    template_text = payload["email_template_text"]
    mapping = payload["mapping"]
    output_dir = Path(payload["output_dir"])
    output_dir.mkdir(parents=True, exist_ok=True)

    placeholders = set(PLACEHOLDER_RE.findall(template_text))
    generated_rows = []
    skipped_rows = []
    drafts = []

    for row_number in range(data_start_row, worksheet.max_row + 1):
        row_values = build_row_values(worksheet, row_number, mapping)
        missing = sorted(
            placeholder for placeholder in placeholders if not row_values.get(placeholder, "")
        )

        if missing:
            skipped_rows.append((row_number, missing))
            continue

        rendered = render_template(template_text, row_values)
        drafts.append((row_number, rendered))
        generated_rows.append(row_number)

    combined_email_path = output_dir / "email_drafts.txt"
    report_path = output_dir / "generation_report.txt"

    write_email_drafts(combined_email_path, drafts)
    write_report(report_path, payload, placeholders, generated_rows, skipped_rows)

    print(f"Generated {len(generated_rows)} emails.")
    print(f"Skipped {len(skipped_rows)} rows.")
    print(f"Combined email drafts file: {combined_email_path}")
    print(f"Report: {report_path}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise
